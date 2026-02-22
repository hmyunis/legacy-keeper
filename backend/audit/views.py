import json
from io import BytesIO
from datetime import datetime, time, timedelta

from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rest_framework import decorators, exceptions, permissions, status, viewsets
from rest_framework.response import Response

from vaults.models import Membership

from .models import AuditLog
from .serializers import AuditLogSerializer

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = AuditLogSerializer
    permission_classes = [permissions.IsAuthenticated]

    def _vault_id(self):
        return (self.request.query_params.get('vault') or '').strip()

    def _require_admin_access(self, vault_id):
        is_admin = Membership.objects.filter(
            vault_id=vault_id,
            user=self.request.user,
            role=Membership.Roles.ADMIN,
            is_active=True,
        ).exists()
        if not is_admin:
            raise exceptions.PermissionDenied("Vault admin access required")

    def _parse_timestamp(self, raw_value, end_of_day=False):
        if not raw_value:
            return None

        parsed_dt = parse_datetime(raw_value)
        if not parsed_dt:
            parsed_date = parse_date(raw_value)
            if parsed_date:
                parsed_dt = datetime.combine(parsed_date, time.max if end_of_day else time.min)

        if not parsed_dt:
            return None

        if timezone.is_naive(parsed_dt):
            parsed_dt = timezone.make_aware(parsed_dt, timezone.get_current_timezone())

        return parsed_dt

    def _apply_filters(self, queryset):
        query_params = self.request.query_params

        category = (query_params.get('category') or '').strip().lower()
        if category == 'uploads':
            queryset = queryset.filter(content_type__model='mediaitem')
        elif category == 'access':
            queryset = queryset.filter(
                Q(action__in=(AuditLog.Action.LOGIN, AuditLog.Action.ACCESS))
                | Q(content_type__model__in=('membership', 'invite'))
            )
        elif category == 'system':
            queryset = queryset.filter(actor__isnull=True)
        elif category == 'management':
            queryset = queryset.exclude(content_type__model='mediaitem')

        timeframe = (query_params.get('timeframe') or '').strip().upper()
        if timeframe in {'DAY', 'WEEK', 'MONTH'}:
            now = timezone.now()
            threshold = now - timedelta(days=1 if timeframe == 'DAY' else 7 if timeframe == 'WEEK' else 30)
            queryset = queryset.filter(timestamp__gte=threshold)

        timestamp_after = self._parse_timestamp(
            query_params.get('timestampAfter')
            or query_params.get('timestamp_after')
            or query_params.get('dateFrom')
        )
        if timestamp_after:
            queryset = queryset.filter(timestamp__gte=timestamp_after)

        timestamp_before = self._parse_timestamp(
            query_params.get('timestampBefore')
            or query_params.get('timestamp_before')
            or query_params.get('dateTo'),
            end_of_day=True,
        )
        if timestamp_before:
            queryset = queryset.filter(timestamp__lte=timestamp_before)

        action = (query_params.get('action') or '').strip().upper()
        if action in dict(AuditLog.Action.choices):
            queryset = queryset.filter(action=action)

        actor_query = (query_params.get('actor') or '').strip()
        if actor_query:
            queryset = queryset.filter(
                Q(actor__full_name__icontains=actor_query) | Q(actor__email__icontains=actor_query)
            )

        target_type = (query_params.get('targetType') or query_params.get('target_type') or '').strip().lower()
        if target_type:
            queryset = queryset.filter(content_type__model__icontains=target_type)

        text_query = (query_params.get('q') or query_params.get('search') or '').strip()
        if text_query:
            queryset = queryset.filter(
                Q(actor__full_name__icontains=text_query)
                | Q(actor__email__icontains=text_query)
                | Q(action__icontains=text_query)
                | Q(content_type__model__icontains=text_query)
            )

        return queryset

    def get_queryset(self):
        """
        Filter logs by the vault ID passed in the query param.
        Example: /api/audit/?vault={uuid}
        """
        vault_pk = self._vault_id()
        if not vault_pk:
            return AuditLog.objects.none()

        self._require_admin_access(vault_pk)
        queryset = AuditLog.objects.filter(vault_id=vault_pk).select_related('actor', 'content_type')
        return self._apply_filters(queryset)

    @decorators.action(detail=False, methods=['get'])
    def export(self, request, *args, **kwargs):
        vault_pk = self._vault_id()
        if not vault_pk:
            return Response({"error": "vault query parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        queryset = self.get_queryset()

        filename = f'audit-log-{vault_pk}-{timezone.now().date().isoformat()}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Audit Logs'

        headers = ['Audit ID', 'Timestamp', 'Actor', 'Action', 'Resource Type', 'Target ID', 'Details']
        worksheet.append(headers)
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = f'A1:G1'

        column_widths = {
            'A': 40,
            'B': 24,
            'C': 24,
            'D': 14,
            'E': 22,
            'F': 40,
            'G': 88,
        }
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        header_fill = PatternFill(fill_type='solid', fgColor='1E3A8A')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1'),
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        odd_row_fill = PatternFill(fill_type='solid', fgColor='F8FAFC')
        even_row_fill = PatternFill(fill_type='solid', fgColor='FFFFFF')
        body_alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
        details_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        for row_index, log in enumerate(queryset.iterator(chunk_size=500), start=2):
            actor_name = log.actor.full_name if log.actor else 'System'
            timestamp_value = timezone.localtime(log.timestamp).strftime('%Y-%m-%d %H:%M:%S %Z')
            details_value = json.dumps(log.changes or {}, ensure_ascii=False)

            worksheet.append(
                [
                    str(log.id),
                    timestamp_value,
                    actor_name,
                    log.action,
                    log.content_type.model,
                    str(log.object_id),
                    details_value,
                ]
            )

            row_fill = odd_row_fill if row_index % 2 == 1 else even_row_fill
            for cell in worksheet[row_index]:
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = details_alignment if cell.column_letter == 'G' else body_alignment

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        return response
